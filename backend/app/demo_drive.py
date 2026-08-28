"""The connected drive: a folder of real files the agent actually reads.

Only the transport is simulated — instead of the Drive API there is a local
folder. Everything else is genuine: the workbook exists, the parser walks its
cells, and every figure the agent cites carries the coordinates it was really
found at. Change a number in the sheet and the next pull cites the new value.

Swapping in real Drive is `files().get_media` + the same parser.
"""
import csv
import logging
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .schemas import SourceFact

log = logging.getLogger(__name__)

DRIVE_DIR = Path(__file__).resolve().parent.parent / "demo_drive"


def ensure_seed_files() -> None:
    """Create the demo company's files once. Idempotent, safe at import."""
    DRIVE_DIR.mkdir(exist_ok=True)
    xlsx = DRIVE_DIR / "revenue.xlsx"
    if not xlsx.exists():
        wb = Workbook()
        s = wb.active
        s.title = "Summary"
        rows = [
            ("Metric", "Q2 2026"),
            ("ARR", "$8.4M"),
            ("ARR growth (YoY)", "212%"),
            ("New ARR added (Q2)", "$1.9M"),
            ("Burn multiple", "1.1x"),
            ("Runway at current burn", "14 months"),
            ("Magic number", "1.4"),
            ("CAC payback", "11 months"),
            ("Median contract length", "24 months"),
            ("Expansion revenue share", "38%"),
            ("Churned logos (Q2)", "1"),
            ("Net revenue retention", "142%"),
            ("Logo retention", "97%"),
        ]
        for r in rows:
            s.append(r)
        pl = wb.create_sheet("P&L")
        pl["C2"] = "P&L line"; pl["D2"] = "Q2 2026"
        pl["C3"] = "Revenue (quarterly)"; pl["D3"] = "$2.3M"
        pl["C4"] = "COGS"; pl["D4"] = "$0.44M"
        pl["C5"] = "Infra spend"; pl["D5"] = "$0.31M"
        pl["C6"] = "Support cost"; pl["D6"] = "$0.13M"
        pl["C7"] = "Gross margin"; pl["D7"] = "81%"
        pl["C8"] = "Opex"; pl["D8"] = "$2.1M"
        acc = wb.create_sheet("Accounts")
        acc.append(("Segment", "Count", "Net retention"))
        acc.append(("Enterprise logos", 38, "151%"))
        acc.append(("Mid-market logos", 74, "128%"))
        acc.append(("Self-serve teams", 412, "104%"))
        acc.append(("Fortune 500 of the 38", 9, ""))
        wb.save(xlsx)
    beta = DRIVE_DIR / "beta.xlsx"
    if not beta.exists():
        wb = Workbook()
        b = wb.active
        b.title = "Beta"
        for row in [
            ("Metric", "Value"),
            ("Beta teams live", 62),
            ("Weekly active usage", "87%"),
            ("Median time-to-resolution (before)", "41 min"),
            ("Median time-to-resolution (with Autopilot)", "24 min"),
            ("Resolution time improvement", "41%"),
            ("Incidents auto-remediated end to end", "63%"),
            ("Sev-1 regressions caused by Autopilot", 0),
            ("Beta NPS", 58),
            ("Error budget consumed (beta fleet)", "31%"),
            ("Beta accounts requesting approval mode", 11),
        ]:
            b.append(row)
        wb.save(beta)
    cohorts = DRIVE_DIR / "cohorts.csv"
    if not cohorts.exists():
        with cohorts.open("w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Cohort", "Month 3", "Month 6", "Month 12", "Net retention"])
            w.writerow(["2024-H2", "96%", "104%", "121%", "134%"])
            w.writerow(["2025-H1", "97%", "108%", "126%", "141%"])
            w.writerow(["2025-H2", "98%", "111%", "129%", "146%"])
            w.writerow(["2026-H1", "99%", "114%", "—", "—"])


def _match_file(hint: str) -> Path | None:
    """Loose match: "the revenue sheet" → revenue.xlsx."""
    ensure_seed_files()
    files = [p for p in DRIVE_DIR.iterdir() if p.is_file() and not p.name.startswith(".")]
    tokens = re.findall(r"[a-z0-9]+", (hint or "").lower())
    best, score = None, 0
    for p in files:
        hits = sum(1 for t in tokens if t in p.name.lower())
        if hits > score:
            best, score = p, hits
    return best or (files[0] if files else None)


def read_file(hint: str) -> tuple[str, list[SourceFact]]:
    """Open the best-matching file and pull out labelled figures with the
    coordinates they were actually found at."""
    path = _match_file(hint)
    if path is None:
        return "", []
    if path.suffix == ".xlsx":
        return path.name, _read_xlsx(path)
    if path.suffix == ".csv":
        return path.name, _read_csv(path)
    return path.name, []


def _looks_numeric(v) -> bool:
    return bool(re.search(r"\d", str(v))) if v is not None else False


def _read_xlsx(path: Path) -> list[SourceFact]:
    wb = load_workbook(path, data_only=True)
    facts: list[SourceFact] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for a, b in zip(row, row[1:]):
                label, value = a.value, b.value
                if (
                    isinstance(label, str) and label.strip()
                    and not label.lower().startswith(("metric", "p&l line", "segment"))
                    and _looks_numeric(value)
                ):
                    facts.append(SourceFact(
                        fact=label.strip(),
                        value=str(value).strip(),
                        source_doc=path.name,
                        source_ref=f"{path.name} · {sheet.title}!{b.coordinate}",
                    ))
    return facts[:14]


def _read_csv(path: Path) -> list[SourceFact]:
    with path.open() as f:
        rows = list(csv.reader(f))
    if len(rows) < 2:
        return []
    header, facts = rows[0], []
    # The most recent complete cohort is the story; cite its actual row number.
    for idx in range(len(rows) - 1, 0, -1):
        if all(c and c != "—" for c in rows[idx]):
            for col in range(1, len(header)):
                facts.append(SourceFact(
                    fact=f"{rows[idx][0]} cohort — {header[col]}",
                    value=rows[idx][col],
                    source_doc=path.name,
                    source_ref=f"{path.name} · row {idx + 1}",
                ))
            break
    return facts
